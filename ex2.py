import openpyxl
import os

# Lista de archivos Excel a procesar
files_to_process = [
    'Spot_Price_List_cuartos.xlsx',
    'aFRRup_HW09.xlsx',
    'Heat_Demand.xlsx',
    'aFRRdown_DM33.xlsx',
    'aFRRdown_DM34.xlsx',
    'aFRRdown_HW09.xlsx',
    'aFRRup_DM33.xlsx',
    'aFRRup_DM34.xlsx'
]

# Crear la carpeta Resultados si no existe
if not os.path.exists('Resultados'):
    os.makedirs('Resultados')
    print("Carpeta 'Resultados' creada.")

def transpose_excel(input_file):
    """
    Transpone un archivo Excel según las especificaciones:
    - La última columna se convierte en la primera fila
    - La primera fila se convierte en la primera columna (desde fila 2)
    - Los valores se transponen correctamente
    """
    try:
        # Abrir el archivo Excel original
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active
        
        # Leer los datos
        data = []
        for row in sheet.rows:
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        # Crear la nueva matriz transpuesta
        new_data = []
        
        # PASO 1: La última columna del original se vuelve la primera fila del nuevo
        last_column = [row[-1] if row else None for row in data]
        new_data.append(last_column)
        
        # PASO 2: La primera fila del original se vuelve la primera columna del nuevo
        first_row = data[0][:-1] if data else []  # Sin el último elemento
        
        # PASO 3: Agregar los valores transpuestos
        for i in range(1, len(data)):
            row = data[i]
            if row:
                row_without_last = row[:-1]  # Sin el último elemento
                
                for j in range(len(row_without_last)):
                    # Si la columna no existe, crearla
                    if len(new_data) <= j + 1:
                        new_data.append([])
                    
                    # Agregar el valor a la columna correspondiente
                    new_data[j + 1].append(row_without_last[j])
        
        # Insertar la primera fila original como primera columna
        for i in range(len(first_row)):
            if len(new_data) > i + 1:
                new_data[i + 1].insert(0, first_row[i])
        
        # Crear un nuevo libro Excel
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Transposed"
        
        # Escribir los datos transpuestos
        for row_idx, row_data in enumerate(new_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                new_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar el nuevo archivo en la carpeta Resultados
        output_filename = os.path.join('Resultados', f"{os.path.splitext(input_file)[0]}_transposed.xlsx")
        new_workbook.save(output_filename)
        
        return True, f"Archivo guardado: {output_filename}"
    
    except Exception as e:
        return False, f"Error al procesar {input_file}: {str(e)}"

# Procesar cada archivo
results = []
for file in files_to_process:
    if os.path.exists(file):
        success, message = transpose_excel(file)
        results.append((file, success, message))
        print(message)
    else:
        results.append((file, False, f"Archivo no encontrado: {file}"))
        print(f"Archivo no encontrado: {file}")

# Resumen final
print("\n" + "="*50)
print("RESUMEN DE PROCESAMIENTO")
print("="*50)

for file, success, message in results:
    status = "✓" if success else "✗"
    print(f"{status} {file}: {message}")

print("\nTodos los archivos transpuestos se han guardado en la carpeta 'Resultados'")