import openpyxl
import os

# Obtener la ruta actual del script
script_dir = os.path.dirname(os.path.abspath(__file__))
print(f"Ruta actual del script: {script_dir}")

# Lista de archivos Excel a procesar
files_to_process = [
    'Spot_Price_List_cuartos.xlsx',
    'aFRRup_HW09.xlsx',
    'Heat_Demand.xlsx',
    'aFRRdown_DM33.xlsx',
    'aFRRdown_DM34.xlsx',
    'aFRRdown_HW09.xlsx',
    'aFRRup_DM33.xlsx',
    'aFRRup_DM34.xlsx'
]

# Buscar archivos Excel en la ruta actual
def find_excel_files(directory):
    """Encuentra todos los archivos Excel en el directorio especificado"""
    found_files = []
    for file in files_to_process:
        file_path = os.path.join(directory, file)
        if os.path.exists(file_path):
            found_files.append(file_path)
    return found_files

# Obtener archivos Excel disponibles
excel_files = find_excel_files(script_dir)
print(f"\nArchivos Excel encontrados: {len(excel_files)}")
for file in excel_files:
    print(f"  - {os.path.basename(file)}")

# Crear la carpeta Resultados en la ruta actual
results_dir = os.path.join(script_dir, 'Resultados')
os.makedirs(results_dir, exist_ok=True)
print(f"\nCarpeta 'Resultados' en: {results_dir}")

def compress_data(transposed_data):
    """
    Comprime los datos agrupando cada 4 filas en una sola hora
    También convierte las horas a formato simplificado (1-24)
    """
    if not transposed_data or len(transposed_data) <= 1:
        return transposed_data
    
    compressed_data = []
    
    # La primera fila (fechas) se mantiene igual
    compressed_data.append(transposed_data[0])
    
    # Procesar el resto de datos agrupando cada 4 filas
    i = 1
    hour_counter = 1
    while i < len(transposed_data):
        # Tomar solo cada 4ta fila para representar la hora completa
        if i + 3 < len(transposed_data):
            # Copiar la fila original
            row_to_add = transposed_data[i].copy()
            # Reemplazar la primera columna (hora) con el número de hora simplificado
            row_to_add[0] = hour_counter
            compressed_data.append(row_to_add)
            hour_counter += 1
            i += 4
        else:
            # Si quedan menos de 4 filas al final, agregar lo que queda
            row_to_add = transposed_data[i].copy()
            row_to_add[0] = hour_counter
            compressed_data.append(row_to_add)
            hour_counter += 1
            i += 1
    
    return compressed_data

def transpose_excel(input_file):
    """
    Transpone un archivo Excel según las especificaciones:
    - La última columna se convierte en la primera fila
    - La primera fila se convierte en la primera columna (desde fila 2)
    - Los valores se transponen correctamente
    - Comprime los datos agrupando cada 4 filas en una hora
    """
    try:
        file_name = os.path.basename(input_file)
        print(f"\nProcesando: {file_name}")
        
        # Abrir el archivo Excel original
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active
        
        # Leer los datos
        data = []
        for row in sheet.rows:
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        # Crear la nueva matriz transpuesta
        new_data = []
        
        # PASO 1: La última columna del original se vuelve la primera fila del nuevo
        last_column = [row[-1] if row else None for row in data]
        new_data.append(last_column)
        
        # PASO 2: La primera fila del original se vuelve la primera columna del nuevo
        first_row = data[0][:-1] if data else []  # Sin el último elemento
        
        # PASO 3: Agregar los valores transpuestos
        for i in range(1, len(data)):
            row = data[i]
            if row:
                row_without_last = row[:-1]  # Sin el último elemento
                
                for j in range(len(row_without_last)):
                    # Si la columna no existe, crearla
                    if len(new_data) <= j + 1:
                        new_data.append([])
                    
                    # Agregar el valor a la columna correspondiente
                    new_data[j + 1].append(row_without_last[j])
        
        # Insertar la primera fila original como primera columna
        for i in range(len(first_row)):
            if len(new_data) > i + 1:
                new_data[i + 1].insert(0, first_row[i])
        
        # Comprimir los datos (NUEVO PASO)
        compressed_data = compress_data(new_data)
        
        # Crear un nuevo libro Excel
        new_workbook = openpyxl.Workbook()
        
        # Crear hoja transpuesta original
        original_sheet = new_workbook.active
        original_sheet.title = "Transposed"
        
        # Escribir los datos transpuestos (sin comprimir)
        for row_idx, row_data in enumerate(new_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                original_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Crear hoja comprimida
        compressed_sheet = new_workbook.create_sheet(title="Compressed")
        
        # Escribir los datos comprimidos
        for row_idx, row_data in enumerate(compressed_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                compressed_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar el nuevo archivo en la carpeta Resultados
        base_name = os.path.splitext(file_name)[0]
        output_filename = os.path.join(results_dir, f"{base_name}_transposed_compressed.xlsx")
        new_workbook.save(output_filename)
        
        # Mostrar estadísticas
        print(f"  Filas originales (transpuestas): {len(new_data)}")
        print(f"  Filas comprimidas: {len(compressed_data)}")
        
        return True, f"Archivo guardado: {output_filename}"
    
    except Exception as e:
        return False, f"Error al procesar {file_name}: {str(e)}"

# Procesar cada archivo encontrado
results = []
if excel_files:
    print("\n" + "="*50)
    print("COMENZANDO TRANSPOSICIÓN Y COMPRESIÓN")
    print("="*50)
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        success, message = transpose_excel(file_path)
        results.append((file_name, success, message))
        print(message)
else:
    print("\nNo se encontraron archivos Excel para procesar en la ruta actual.")

# Resumen final
print("\n" + "="*50)
print("RESUMEN DE PROCESAMIENTO")
print("="*50)

for file, success, message in results:
    status = "✓" if success else "✗"
    print(f"{status} {file}: {message}")

print(f"\nTodos los archivos transpuestos y comprimidos se han guardado en la carpeta 'Resultados'")
print("\nCada archivo Excel ahora contiene:")
print("  - Hoja 'Transposed': Datos transpuestos originales")
print("  - Hoja 'Compressed': Datos comprimidos (4 filas -> 1 fila)")